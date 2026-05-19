"""
Generates the staffing schedule xlsx from output/01_extracted.json.

Input : output/01_extracted.json
Output: output/{client}_{event_name}.xlsx
"""
import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("오류: openpyxl이 설치되지 않았습니다. pip install openpyxl 실행 후 재시도하세요.")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).parents[4]
OUTPUT_DIR = PROJECT_ROOT / "output"

DAYS_KO = ["월", "화", "수", "목", "금", "토", "일"]

# Column layout constants
COL_GROUP = 1   # 구분 (역할 그룹)
COL_ROLE = 2    # 역할명
COL_NAME = 3    # 이름 (blank)
DATE_COL_START = 4  # first date column


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sanitize(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|\[\]]', "_", name)


def date_range(start: str, end: str):
    s = datetime.strptime(start, "%Y-%m-%d")
    e = datetime.strptime(end, "%Y-%m-%d")
    d = s
    while d <= e:
        yield d
        d += timedelta(days=1)


def collect_dates(data: dict) -> list[datetime]:
    """Return sorted list of all dates that appear in schedule or event range."""
    dates: set[str] = set()

    for item in data.get("schedule", []):
        dates.add(item["date"])

    for item in data.get("additional_schedules", []):
        d = item.get("date", "")
        if re.match(r"\d{4}-\d{2}-\d{2}", d):
            dates.add(d)

    start = data.get("start_date", "")
    end = data.get("end_date", "")
    if start and end:
        for dt in date_range(start, end):
            dates.add(dt.strftime("%Y-%m-%d"))

    return sorted([datetime.strptime(d, "%Y-%m-%d") for d in dates])


def build_role_rows(data: dict, all_dates: list[datetime]) -> list[dict]:
    """
    Build ordered list of role rows.
    Each row: {role, group, times: {date_str -> time_str}}
    Rows are expanded by count (팀장 count=2 → 2 rows).
    """
    date_strs = {dt.strftime("%Y-%m-%d") for dt in all_dates}

    # Collect unique roles in encounter order, max count per role
    role_order: list[str] = []
    role_max: dict[str, int] = {}
    role_times: dict[tuple, str] = {}  # (date, role) -> time string

    for sched in data.get("schedule", []):
        date = sched["date"]
        if date not in date_strs:
            continue
        for staff in sched.get("staff", []):
            role = staff["role"]
            count = int(staff.get("count", 1))
            start = staff.get("start", "")
            end = staff.get("end", "")
            time_str = f"{start} - {end}" if start and end else ""

            if role not in role_max:
                role_order.append(role)
                role_max[role] = 0
            role_max[role] = max(role_max[role], count)
            role_times[(date, role)] = time_str

    rows = []
    for role in role_order:
        for i in range(role_max[role]):
            times = {}
            for (date, r), t in role_times.items():
                if r == role:
                    times[date] = t
            rows.append({"role": role, "index": i, "times": times})

    return rows


def derive_group(role: str) -> str:
    """Derive a display group label from role string for column A."""
    role_lower = role.lower()
    if "팀장" in role or "sv" in role_lower or "슈퍼" in role_lower:
        return "SV"
    if "세팅" in role or "설치" in role or "철수" in role:
        return "세팅/철수"
    if "영어" in role:
        return "영어"
    if "중국어" in role:
        return "중국어"
    if "일반" in role or "재고" in role or "캐셔" in role or "결제" in role:
        return "일반"
    return role


def get_label_for_date(data: dict, date_str: str) -> str:
    """Return label for special dates (세팅/오픈/철수 etc.)"""
    for item in data.get("schedule", []):
        if item["date"] == date_str and item.get("label"):
            return item["label"]
    return ""


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def header_fill():
    return PatternFill("solid", fgColor="D9E1F2")


def label_fill():
    return PatternFill("solid", fgColor="E2EFDA")


def apply_header_style(cell):
    cell.fill = header_fill()
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def apply_label_style(cell):
    cell.fill = label_fill()
    cell.font = Font(bold=False)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()


def apply_data_style(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------

def build_sheet(ws, data: dict) -> None:
    all_dates = collect_dates(data)
    if not all_dates:
        print("경고: schedule 데이터가 없어 날짜 열을 생성할 수 없습니다.")
        return

    phone_col = DATE_COL_START + len(all_dates)

    # ── Row 1: date numbers ──────────────────────────────────────────────────
    for col_label, text in [
        (COL_GROUP, "구분"),
        (COL_ROLE, "역할"),
        (COL_NAME, "이름"),
        (phone_col, "연락처"),
    ]:
        c = ws.cell(row=1, column=col_label, value=text)
        apply_header_style(c)

    date_col_map: dict[str, int] = {}
    for i, dt in enumerate(all_dates):
        col = DATE_COL_START + i
        date_str = dt.strftime("%Y-%m-%d")
        label = get_label_for_date(data, date_str)
        display = f"{dt.day}일"
        if label:
            display += f"\n[{label}]"
        c = ws.cell(row=1, column=col, value=display)
        apply_header_style(c)
        date_col_map[date_str] = col

    # ── Row 2: day of week ───────────────────────────────────────────────────
    for col in [COL_GROUP, COL_ROLE, COL_NAME, phone_col]:
        ws.cell(row=2, column=col).border = thin_border()

    for i, dt in enumerate(all_dates):
        col = DATE_COL_START + i
        c = ws.cell(row=2, column=col, value=DAYS_KO[dt.weekday()])
        apply_header_style(c)

    # ── Staff rows ───────────────────────────────────────────────────────────
    rows = build_role_rows(data, all_dates)

    # Group label merging: track (group, start_row) per group
    group_start: dict[str, int] = {}
    group_last: dict[str, int] = {}

    current_row = 3
    prev_group = None

    for row_data in rows:
        role = row_data["role"]
        group = derive_group(role)
        row = current_row

        # Group column
        if group != prev_group:
            c = ws.cell(row=row, column=COL_GROUP, value=group)
            apply_label_style(c)
            group_start[group] = row
        else:
            ws.cell(row=row, column=COL_GROUP).border = thin_border()
        group_last[group] = row
        prev_group = group

        # Role column
        c = ws.cell(row=row, column=COL_ROLE, value=role)
        apply_label_style(c)

        # Name column (blank)
        c = ws.cell(row=row, column=COL_NAME, value="")
        apply_data_style(c)

        # Phone column (blank)
        c = ws.cell(row=row, column=phone_col, value="")
        apply_data_style(c)

        # Date columns
        for date_str, col in date_col_map.items():
            time_str = row_data["times"].get(date_str, "")
            c = ws.cell(row=row, column=col, value=time_str)
            apply_data_style(c)

        current_row += 1

    # Merge group cells
    for group, start_row in group_start.items():
        end_row = group_last[group]
        if end_row > start_row:
            ws.merge_cells(
                start_row=start_row, start_column=COL_GROUP,
                end_row=end_row, end_column=COL_GROUP
            )
            merged_cell = ws.cell(row=start_row, column=COL_GROUP)
            merged_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_GROUP)].width = 10
    ws.column_dimensions[get_column_letter(COL_ROLE)].width = 18
    ws.column_dimensions[get_column_letter(COL_NAME)].width = 10
    ws.column_dimensions[get_column_letter(phone_col)].width = 14
    for col in range(DATE_COL_START, phone_col):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # Freeze panes: freeze date headers and left columns
    ws.freeze_panes = ws.cell(row=3, column=DATE_COL_START)


# ---------------------------------------------------------------------------
# Amendment handling
# ---------------------------------------------------------------------------

def handle_amendment(client: str, event_name: str) -> Path | None:
    """Find existing xlsx and mark first sheet as _수정전. Return path or None."""
    pattern = f"{sanitize(client)}_{sanitize(event_name)}"
    for f in sorted(OUTPUT_DIR.glob("*.xlsx")):
        if "_수정전" in f.stem:
            continue
        if pattern in f.stem:
            wb = load_workbook(f)
            for ws in wb.worksheets:
                if "_수정전" not in ws.title:
                    ws.title = (ws.title + "_수정전")[:31]
            wb.save(f)
            return f
    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    json_path = OUTPUT_DIR / "01_extracted.json"
    if not json_path.exists():
        print(f"오류: {json_path} 파일이 없습니다. [1] 단계를 먼저 완료하세요.")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    client = data.get("client", "거래처")
    event_name = data.get("event_name", "행사")
    is_amendment = data.get("is_amendment", False)

    # Handle amendment
    if is_amendment:
        existing = handle_amendment(client, event_name)
        if existing:
            print(f"기존 파일 발견: {existing.name}")
            print(f"  → 시트 탭명에 '_수정전' 추가 완료")
        else:
            print("기존 파일을 찾지 못했습니다. 신규 파일로 생성합니다.")

    # Build new workbook
    wb = Workbook()
    ws = wb.active
    sheet_name = sanitize(f"{client}_{event_name}")[:31]
    ws.title = sheet_name

    build_sheet(ws, data)

    # Determine output path
    out_name = f"{sanitize(client)}_{sanitize(event_name)}.xlsx"
    out_path = OUTPUT_DIR / out_name

    if out_path.exists() and not is_amendment:
        stamp = datetime.now().strftime("%Y%m%d")
        out_name = f"{sanitize(client)}_{sanitize(event_name)}_{stamp}.xlsx"
        out_path = OUTPUT_DIR / out_name

    wb.save(out_path)
    print(f"파일 생성 완료: {out_path}")
    print(f"  → 이름·연락처 셀은 빈 상태입니다. 섭외팀에서 직접 입력해 주세요.")


if __name__ == "__main__":
    main()
